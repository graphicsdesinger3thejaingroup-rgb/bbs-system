"""
Excel file generator using OpenPyXL.
Creates a workbook with two sheets:
  - BBS_Data
  - Final_Order
Plus a small summary header.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
import os
from datetime import datetime


THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, size=10, color="555555")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 38)


def _write_table(ws, start_row: int, headers, rows):
    # header
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    # body
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
    return start_row + 1 + len(rows)


def generate_excel(result: Dict[str, Any], out_dir: str, project_name: str = "BBS_Report") -> str:
    """Generate Excel and return the absolute file path."""
    os.makedirs(out_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{project_name}_{timestamp}.xlsx"
    path = os.path.join(out_dir, filename)

    wb = Workbook()

    # ---------- Sheet 1: BBS_Data ----------
    ws1 = wb.active
    ws1.title = "BBS_Data"

    ws1["A1"] = "Bar Bending Schedule (BBS)"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:J1")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')} | "\
                f"Element: {result['summary']['element_type'].title()} | Standard: IS 456 / IS 2502"
    ws1["A2"].font = SUB_FONT
    ws1.merge_cells("A2:J2")

    headers = ["Bar Mark", "Element", "Description", "Dia (mm)", "Shape",
               "Cutting Length (mm)", "Quantity", "Total Length (m)",
               "Unit Wt (kg/m)", "Total Weight (kg)"]

    body = [
        [r["bar_mark"], r.get("element_type", "-").title(), r["description"],
         r["dia_mm"], r["shape"],
         r["cutting_length_mm"], r["quantity"], r["total_length_m"],
         r["unit_weight_kg_m"], r["total_weight_kg"]]
        for r in result["bbs"]
    ]

    next_row = _write_table(ws1, start_row=4, headers=headers, rows=body)

    # totals
    ws1.cell(row=next_row, column=1, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=next_row, column=8, value=result["summary"]["total_length_m"]).font = Font(bold=True)
    ws1.cell(row=next_row, column=10, value=result["summary"]["total_weight_kg_net"]).font = Font(bold=True)
    for j in range(1, 11):
        ws1.cell(row=next_row, column=j).fill = PatternFill("solid", fgColor="DCE6F1")
        ws1.cell(row=next_row, column=j).border = BORDER

    # Verification stamp row
    if "verification" in result:
        v = result["verification"]
        next_row += 2
        ws1.cell(row=next_row, column=1,
                 value="✓ Engineering Verified" if v["passed"]
                       else "✗ VERIFICATION FAILED").font = Font(
                           bold=True,
                           color="2E7D32" if v["passed"] else "C62828")
        ws1.cell(row=next_row, column=2,
                 value=f"({v.get('method', 'double-check')})").font = SUB_FONT

    _autosize(ws1)

    # ---------- Sheet 2: Final_Order ----------
    ws2 = wb.create_sheet("Final_Order")
    ws2["A1"] = "Final Steel Order — Supplier Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:G1")
    ws2["A2"] = "Includes 5% wastage. Standard rod length = 12 m."
    ws2["A2"].font = SUB_FONT
    ws2.merge_cells("A2:G2")

    headers2 = ["Diameter (mm)", "Net Length (m)", "Net Weight (kg)",
                "Wastage %", "Order Length (m)", "Order Weight (kg)",
                "12m Rods Required"]
    body2 = [
        [r["dia_mm"], r["net_length_m"], r["net_weight_kg"],
         r["wastage_pct"], r["order_length_m"], r["order_weight_kg"],
         r["rods_12m_required"]]
        for r in result["final_order"]
    ]

    nr = _write_table(ws2, start_row=4, headers=headers2, rows=body2)

    ws2.cell(row=nr, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws2.cell(row=nr, column=5, value=result["summary"]["total_order_length_m"]).font = Font(bold=True)
    ws2.cell(row=nr, column=6, value=result["summary"]["total_weight_kg_with_wastage"]).font = Font(bold=True)
    for j in range(1, 8):
        ws2.cell(row=nr, column=j).fill = PatternFill("solid", fgColor="DCE6F1")
        ws2.cell(row=nr, column=j).border = BORDER

    _autosize(ws2)

    wb.save(path)
    return path
